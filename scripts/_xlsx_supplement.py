"""讀取 data/timetables/完整時刻表.xlsx 的兩張補充表。

- LINEBOT_JSON sheet：A4 起為跨格切的 JSON 字串（每格上限 30,000 字），需串回再 parse。
  提供 driver_route / conductor_route 補 consist 的 crew_mech / crew_ops。
- 車次總覽 sheet：第 3 列起為資料，欄位含車次與「備註」（如「逢週五、日行駛」「團體列車」）。

公開 API：
- load_crew_supplement(xlsx_path) -> dict[train_no_str, dict]
- load_train_notes(xlsx_path)     -> dict[train_no_str, str]
- DEFAULT_PATH                    -> Path 指向 data/timetables/完整時刻表.xlsx
"""

from __future__ import annotations

import json
from pathlib import Path

DEFAULT_PATH = Path(__file__).parent.parent / "data" / "timetables" / "完整時刻表.xlsx"


def _load_workbook(xlsx_path: Path):
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，請執行 pip install openpyxl") from e
    return openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)


def load_crew_supplement(xlsx_path: Path = DEFAULT_PATH) -> dict[str, dict]:
    """從 LINEBOT_JSON sheet 抽乘務補充資料。

    回傳 {train_no: {driver_route, conductor_route, train_type, depot, route}}
    train_no 永遠為字串（"1", "1004A"）。
    """
    if not xlsx_path.exists():
        return {}
    wb = _load_workbook(xlsx_path)
    if "LINEBOT_JSON" not in wb.sheetnames:
        return {}
    ws = wb["LINEBOT_JSON"]
    parts: list[str] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        for cell in row:
            if cell:
                parts.append(str(cell))
    if not parts:
        return {}
    raw = json.loads("".join(parts))
    return {str(k): v for k, v in raw.get("trains", {}).items()}


def load_train_notes(xlsx_path: Path = DEFAULT_PATH) -> dict[str, str]:
    """從 車次總覽 sheet 抽備註欄。

    回傳 {train_no: notes}，僅納入備註非空者。
    train_no 永遠為字串。
    """
    if not xlsx_path.exists():
        return {}
    wb = _load_workbook(xlsx_path)
    if "車次總覽" not in wb.sheetnames:
        return {}
    ws = wb["車次總覽"]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        train_no = str(row[0]).strip()
        notes = (row[10] or "").strip() if len(row) > 10 and row[10] else ""
        if train_no and notes:
            out[train_no] = notes
    return out


if __name__ == "__main__":
    import sys
    crew = load_crew_supplement()
    notes = load_train_notes()
    print(f"crew supplement   : {len(crew)} trains")
    print(f"  with driver_route   : {sum(1 for v in crew.values() if v.get('driver_route'))}")
    print(f"  with conductor_route: {sum(1 for v in crew.values() if v.get('conductor_route'))}")
    print(f"  with depot          : {sum(1 for v in crew.values() if v.get('depot'))}")
    print(f"train notes       : {len(notes)} trains")
    print(f"\nsample crew['1']: {crew.get('1')}")
    print(f"sample notes['21']: {notes.get('21')}")
