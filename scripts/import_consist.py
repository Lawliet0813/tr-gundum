"""
未來編組更新腳本：從 Excel 讀取台鐵最新編組運用表，更新 consist.json。

使用方式：
  python scripts/import_consist.py 新編組表.xlsx
  python scripts/import_consist.py 新編組表.xlsx --date 1150120
  python scripts/import_consist.py 新編組表.xlsx --sheet 2  （指定第幾個 sheet，預設第 1 個）

Excel 欄位對應（預設欄位名稱，若不同可用 --map 調整）：
  車次、車型、歸屬段與運用號碼、區間、值乘區間(機務)、值乘區間(運務)

--map 範例（JSON 格式，舊欄位名 → 標準欄位名）：
  --map '{"Train No":"車次","Type":"車型"}'
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("請先安裝 openpyxl：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

OUTPUT = Path(__file__).parent.parent / "data" / "consist.json"

# 預設欄位名對應（Excel 欄位名 → JSON key）
DEFAULT_COL_MAP = {
    "車次":            "train_no",
    "車型":            "type_name",
    "歸屬段與運用號碼":  "formation",
    "歸屬段運用號碼":   "formation",
    "運用號碼":        "formation",
    "區間":            "route",
    "值乘區間(機務)":   "crew_mech",
    "值乘區間（機務）":  "crew_mech",
    "機務":            "crew_mech",
    "值乘區間(運務)":   "crew_ops",
    "值乘區間（運務）":  "crew_ops",
    "運務":            "crew_ops",
}


def roc_to_ad(roc: str) -> str:
    """民國 YYYMMDD → YYYY-MM-DD"""
    roc = roc.strip()
    if len(roc) == 7 and roc.isdigit():
        return f"{int(roc[:3]) + 1911}-{roc[3:5]}-{roc[5:7]}"
    return roc


def detect_columns(ws) -> dict[str, int]:
    """掃描第一行找欄位對應的 col index（0-based）。"""
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
    col_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in DEFAULT_COL_MAP:
            key = DEFAULT_COL_MAP[h]
            if key not in col_idx:
                col_idx[key] = i
    return col_idx


def main():
    parser = argparse.ArgumentParser(description="從 Excel 更新台鐵編組運用表")
    parser.add_argument("xlsx", help="Excel 檔案路徑（.xlsx）")
    parser.add_argument("--date", default="", help="民國日期，格式 YYYMMDD，例如 1150120")
    parser.add_argument("--sheet", type=int, default=1, help="第幾個 sheet（從 1 開始，預設 1）")
    parser.add_argument("--map", default="{}", help="JSON 格式的欄位名稱對應")
    parser.add_argument("--merge", action="store_true",
                        help="合併模式：只更新 Excel 中存在的車次，其餘保留原資料")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"找不到檔案：{xlsx_path}", file=sys.stderr)
        sys.exit(1)

    # 套用自訂 col map
    extra_map: dict[str, str] = json.loads(args.map)
    col_map = {**DEFAULT_COL_MAP, **extra_map}

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.worksheets[args.sheet - 1]

    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]

    # 建立欄位 index（優先用 extra_map）
    col_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = col_map.get(h)
        if key and key not in col_idx:
            col_idx[key] = i

    if "train_no" not in col_idx:
        print(f"找不到「車次」欄位，偵測到的欄位：{headers}", file=sys.stderr)
        sys.exit(1)

    def cell_val(row, key: str) -> str:
        idx = col_idx.get(key)
        if idx is None:
            return ""
        v = row[idx].value
        return str(v).strip() if v is not None else ""

    # 讀取資料
    new_trains: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2):
        train_no = cell_val(row, "train_no")
        if not train_no or train_no == "None":
            continue
        new_trains[train_no] = {
            "type_name": cell_val(row, "type_name"),
            "formation": cell_val(row, "formation"),
            "route":     cell_val(row, "route"),
            "crew_mech": cell_val(row, "crew_mech"),
            "crew_ops":  cell_val(row, "crew_ops"),
        }

    if not new_trains:
        print("Excel 中未讀到任何資料，請確認欄位名稱", file=sys.stderr)
        sys.exit(1)

    # 合併或覆蓋
    if args.merge and OUTPUT.exists():
        existing = json.loads(OUTPUT.read_text(encoding="utf-8"))
        merged = {**existing.get("trains", {}), **new_trains}
        trains = merged
    else:
        trains = new_trains

    date_roc = args.date or ""
    result = {
        "version":    date_roc,
        "updated_at": roc_to_ad(date_roc) if date_roc else "",
        "source":     str(xlsx_path.name),
        "trains":     trains,
    }

    OUTPUT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"完成：共 {len(trains)} 筆車次資料 → {OUTPUT}")
    if date_roc:
        print(f"版本：民國 {date_roc}（西元 {roc_to_ad(date_roc)}）")
    print("\n欄位對應結果：")
    for key, idx in col_idx.items():
        print(f"  {key:15s} ← 第 {idx+1} 欄「{headers[idx]}」")


if __name__ == "__main__":
    main()
